import tkinter as tk
from functools import partial
import openpyxl
from tkinter import messagebox

def bt_click(botao):
    global nome_planilha_1
    global nome_planilha_2
    if botao == 'bt1':
        if ed.get() != '':
            lb3['text'] = ed.get()+'.xlsx'
            nome_planilha_1 = ed.get() + '.xlsx'
        else:
            lb3['text'] = ' '
            nome_planilha_1 = ''

    if botao == 'bt2':
        if ed2.get() != '':
            lb4['text'] = ed2.get()+'.xlsx'
            nome_planilha_2 = ed2.get()+ '.xlsx'
        else:
            lb4['text'] = ' '
            nome_planilha_2 = ''

def TBanalisada():
    wb = openpyxl.Workbook()
    planilha_analisada = wb.worksheets[0]

    c1 = (f'C:\ExcelPython\{nome_planilha_1}')
    c2 = (f'C:\ExcelPython\{nome_planilha_2}')
    planilha_1 = openpyxl.load_workbook(c1)
    planilha_2 = openpyxl.load_workbook(c2)

    pla1 = planilha_1.active
    pla2 = planilha_2.active

    for x in range(1, pla1.max_row+1):
        janela.update()
        cel_pla1 = pla1.cell(row=x, column= 1)
        val_pla1 = cel_pla1.value
        cel_pla1B = pla1.cell(row=x, column= 2)
        val_pla1B = cel_pla1B.value
        cel_pla1C = pla1.cell(row=x, column= 3)
        val_pla1C = cel_pla1C.value
        cel_pla1D = pla1.cell(row=x, column= 4)
        val_pla1D = cel_pla1D.value
        cel_pla1E = pla1.cell(row=x, column= 5)
        val_pla1E = cel_pla1E.value
        for i in range(1, pla2.max_row+1):
            janela.update()
            cel_pla2 = pla2.cell(row=i, column= 1)
            val_pla2 = cel_pla2.value
            cel_pla2H = pla2.cell(row=i, column= 2)
            val_pla2H = cel_pla2H.value
            cel_pla2I = pla2.cell(row=i, column= 3)
            val_pla2I = cel_pla2I.value
            cel_pla2J = pla2.cell(row=i, column= 4)
            val_pla2J = cel_pla2J.value
            cel_pla2K = pla2.cell(row=i, column= 5)
            val_pla2K = cel_pla2K.value
            if (val_pla2) == (val_pla1):
                planilha_analisada[f'A{planilha_analisada.max_row+1}'] = val_pla1
                planilha_analisada[f'B{planilha_analisada.max_row}'] = val_pla1B
                planilha_analisada[f'C{planilha_analisada.max_row}'] = val_pla1C
                planilha_analisada[f'D{planilha_analisada.max_row}'] = val_pla1D
                planilha_analisada[f'E{planilha_analisada.max_row}'] = val_pla1E
                planilha_analisada[f'G{planilha_analisada.max_row}'] = val_pla2
                planilha_analisada[f'H{planilha_analisada.max_row}'] = val_pla2H
                planilha_analisada[f'I{planilha_analisada.max_row}'] = val_pla2I
                planilha_analisada[f'J{planilha_analisada.max_row}'] = val_pla2J
                planilha_analisada[f'K{planilha_analisada.max_row}'] = val_pla2K

    wb.save('C:\ExcelPython\Tabela_Analisada.xlsx')

def TBGesthos():


    wb = openpyxl.Workbook()
    planilha_analisada_nao_encontrada_gesthos = wb.worksheets[0]

    c1 = (f'C:\ExcelPython\{nome_planilha_1}')
    c2 = (f'C:\ExcelPython\{nome_planilha_2}')
    planilha_111 = openpyxl.load_workbook(c1)
    planilha_222 = openpyxl.load_workbook(c2)

    pla111 = planilha_111.active
    pla222 = planilha_222.active

    for x in range(1, pla111.max_row+1):
        janela.update()
        p = False
        cel_pla111 = pla111.cell(row=x, column= 1)
        val_pla111 = cel_pla111.value
        cel_pla111B = pla111.cell(row=x, column= 2)
        val_pla111B = cel_pla111B.value
        cel_pla111C = pla111.cell(row=x, column= 3)
        val_pla111C = cel_pla111C.value
        cel_pla111D = pla111.cell(row=x, column= 4)
        val_pla111D = cel_pla111D.value
        cel_pla111E = pla111.cell(row=x, column= 5)
        val_pla111E = cel_pla111E.value
        for i in range(1, pla222.max_row+1):
            janela.update()
            cel_pla222 = pla222.cell(row=i, column= 1)
            val_pla222 = cel_pla222.value
            if (val_pla222) == (val_pla111):
                p = True
        if p == False:
            planilha_analisada_nao_encontrada_gesthos[f'A{planilha_analisada_nao_encontrada_gesthos.max_row+1}'] = val_pla111
            planilha_analisada_nao_encontrada_gesthos[f'B{planilha_analisada_nao_encontrada_gesthos.max_row}'] = val_pla111B
            planilha_analisada_nao_encontrada_gesthos[f'C{planilha_analisada_nao_encontrada_gesthos.max_row}'] = val_pla111C
            planilha_analisada_nao_encontrada_gesthos[f'D{planilha_analisada_nao_encontrada_gesthos.max_row}'] = val_pla111D
            planilha_analisada_nao_encontrada_gesthos[f'E{planilha_analisada_nao_encontrada_gesthos.max_row}'] = val_pla111E


    wb.save('C:\ExcelPython\Tabela_Analisada_Nao_Encontrados_Gesthos.xlsx')

def TBConvenio():
    wb = openpyxl.Workbook()
    planilha_analisada_nao_encontrada = wb.worksheets[0]

    c1 = (f'C:\ExcelPython\{nome_planilha_1}')
    c2 = (f'C:\ExcelPython\{nome_planilha_2}')
    planilha_11 = openpyxl.load_workbook(c2)
    planilha_22 = openpyxl.load_workbook(c1)

    pla11 = planilha_11.active
    pla22 = planilha_22.active

    for x in range(1, pla11.max_row+1):
        o = False
        janela.update()
        cel_pla11 = pla11.cell(row=x, column= 1)
        val_pla11 = cel_pla11.value
        cel_pla11B = pla11.cell(row=x, column= 2)
        val_pla11B = cel_pla11B.value
        cel_pla11C = pla11.cell(row=x, column= 3)
        val_pla11C = cel_pla11C.value
        cel_pla11D = pla11.cell(row=x, column= 4)
        val_pla11D = cel_pla11D.value
        cel_pla11E = pla11.cell(row=x, column= 5)
        val_pla11E = cel_pla11E.value
        for i in range(1, pla22.max_row+1):
            janela.update()
            cel_pla22 = pla22.cell(row=i, column= 1)
            val_pla22 = cel_pla22.value
            if (val_pla22) == (val_pla11):
                o = True
        if o == False:
            planilha_analisada_nao_encontrada[f'A{planilha_analisada_nao_encontrada.max_row+1}'] = val_pla11
            planilha_analisada_nao_encontrada[f'B{planilha_analisada_nao_encontrada.max_row}'] = val_pla11B
            planilha_analisada_nao_encontrada[f'C{planilha_analisada_nao_encontrada.max_row}'] = val_pla11C
            planilha_analisada_nao_encontrada[f'D{planilha_analisada_nao_encontrada.max_row}'] = val_pla11D
            planilha_analisada_nao_encontrada[f'E{planilha_analisada_nao_encontrada.max_row}'] = val_pla11E


    wb.save('C:\ExcelPython\Tabela_Analisada_Nao_Encontrados.xlsx')

def execultar():
    if (nome_planilha_1 != '') and (nome_planilha_2 != ''):
        if vAnalisarTabelas.get() != 0 or vTabelaNaoEncontradoConvenio.get() != 0 or vTabelaNaoEncontradoGesthost.get() != 0:
            lb6['text'] = 'Executando... Por favor aguarde'
            try:
                n = 0
                if vAnalisarTabelas.get() == 1:
                    TBanalisada()
                    n += 1
                if vTabelaNaoEncontradoConvenio.get() == 1:
                    TBConvenio()
                    n +=1
                if vTabelaNaoEncontradoGesthost.get() == 1:
                    TBGesthos()
                    n +=1
                if n == 1:
                    messagebox.showinfo("Sucesso","Tabela Gerada")
                else:
                    messagebox.showinfo("Sucesso","Tabelas Geradas")
                lb6['text'] = ''
            except:
                messagebox.showerror("Erro","Planilhas Não Encontradas, verifique os nomes e tente novamente!")
                lb6['text'] = ''
        else:
            messagebox.showerror("Erro","Selecione as planilhas a serem geradas!") 
    else:
       messagebox.showerror("Erro","Insira tabelas válidas!") 


nome_planilha_1 = ''
nome_planilha_2 = ''


janela = tk.Tk()
janela.resizable(0,0)

janela.iconbitmap('C:\ExcelPython\dist\Icone.ico')


janela.title('ExcelPython')
janela.geometry('800x520+400+150')

lb_titulo = tk.Label(janela, text= 'ExcelPython - Análise de Tabelas Excel')
lb_titulo.place(x=286, y=5)

fr_nomes= tk.Frame(janela, borderwidth = 1, relief='solid')
fr_nomes.place(x=10, y=40, width=780, height=200)

fr_tabelas= tk.Frame(janela, borderwidth = 1, relief='solid')
fr_tabelas.place(x=10, y=260, width=780, height=200)

lb = tk.Label(janela, text= 'Nome da Tabela do GestHost: ')
lb.place(x=20, y=110)
lb2 = tk.Label(janela, text= 'Nome da Tabela do Convênio: ')
lb2.place(x=20, y=145)

ed = tk.Entry(janela)
ed.place(x=200, y=111)

ed2 = tk.Entry(janela)
ed2.place(x=200, y=146)

bt1 = tk.Button(janela, width=5, text='OK')
bt1.place(x=350, y=108)
bt1['command'] = partial (bt_click, 'bt1')
bt2= tk.Button(janela, width=5, text='OK')
bt2.place(x=350, y=143)
bt2['command'] = partial (bt_click, 'bt2')

lb3 = tk.Label(janela, text= '')
lb3.place(x=400, y=110)
lb4 = tk.Label(janela, text= '')
lb4.place(x=400, y=145)
lb6 = tk.Label(janela, text='')
lb6.place(x=291, y=470)

vAnalisarTabelas = tk.IntVar()
vTabelaNaoEncontradoGesthost = tk.IntVar()
vTabelaNaoEncontradoConvenio = tk.IntVar()


cb_vAnalisarTabelas = tk.Checkbutton(janela, text='Gerar tabela de comparação dos itens encontrados', variable=vAnalisarTabelas, onvalue=1, offvalue=0)
cb_vAnalisarTabelas.place(x=20, y=320)
cb_vTabelaNaoEncontradoGesthost = tk.Checkbutton(janela, text='Gerar tabela de itens não encontrados no Gesthost', variable=vTabelaNaoEncontradoGesthost, onvalue=1, offvalue=0)
cb_vTabelaNaoEncontradoGesthost.place(x=20, y=350)
cb_vTabelaNaoEncontradoConvenio = tk.Checkbutton(janela, text='Gerar tabela de itens não encontrados no Convênio', variable=vTabelaNaoEncontradoConvenio, onvalue=1, offvalue=0)
cb_vTabelaNaoEncontradoConvenio.place(x=20, y=380)

bt_gerar = tk.Button(janela, width=10, height=2, text='GERAR', command= execultar)
bt_gerar.place(x=650, y=345)

lb5 = tk.Label(janela, text= 'Desenvolvido por: Luiz Fernando')
lb5.place(x=610, y=500)


janela.mainloop()
